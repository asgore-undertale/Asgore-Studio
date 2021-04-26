#استيراد المكتبات
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import QMainWindow, QTextEdit, QCheckBox, QPushButton, QLabel, QMenuBar, QAction, QFileDialog, QRadioButton, QMenu, QMessageBox
from os import path, listdir, mkdir, makedirs, walk
from openpyxl.styles import PatternFill, Alignment, Font
import openpyxl

##استيراد السكربتات
from Parts.Scripts.Delete_Duplicated_lines import DDL
from Parts.Scripts.Un_Freeze_Arabic import Un_Freeze
from Parts.Scripts.Extract_from_text import Extract
from Parts.Scripts.Convert_bytes import convert_bytes
from Parts.Scripts.Reverse_text import Reverse
from Parts.Scripts.Sort_lines import Sort
from Parts.Scripts.Un_Convert import Convert
from Parts.Scripts.HandleHarakat import handle_harakat
from Parts.Scripts.Take_From_Table import Take_From_Table

##الخطوط
labels_font = QtGui.QFont()
labels_font.setPointSize(9)
textbox_font = QtGui.QFont()
textbox_font.setPointSize(12)

#نافذة خيارات التحويل
OptionsWindow_Width = 400
checkbox_size = [OptionsWindow_Width-5, 16]
def pos_y(line_num, Height = checkbox_size[1], Between_every_y = 15):
    return (Between_every_y+checkbox_size[1]) * line_num + (Between_every_y-Height//2) + (Between_every_y//2)

OptionsWindow = QMainWindow()
OptionsWindow.setFixedSize(OptionsWindow_Width, 380)

DDL_check = QCheckBox("حذف الأسطر المكررة", OptionsWindow)
DDL_check.setGeometry(QtCore.QRect(0, pos_y(0), checkbox_size[0], checkbox_size[1]))
DDL_check.setLayoutDirection(QtCore.Qt.RightToLeft)
SSL_check = QCheckBox("ترتيب السطور من الأقصر للأطول", OptionsWindow)
SSL_check.setGeometry(QtCore.QRect(0, pos_y(1), checkbox_size[0], checkbox_size[1]))
SSL_check.setLayoutDirection(QtCore.Qt.RightToLeft)
SLS_check = QCheckBox("ترتيب السطور من الأطول للأقصر", OptionsWindow)
SLS_check.setGeometry(QtCore.QRect(0, pos_y(2), checkbox_size[0], checkbox_size[1]))
SLS_check.setLayoutDirection(QtCore.Qt.RightToLeft)
RA_check = QCheckBox("تجميد النص العربي", OptionsWindow)
RA_check.setGeometry(QtCore.QRect(0, pos_y(3), checkbox_size[0], checkbox_size[1]))
RA_check.setLayoutDirection(QtCore.Qt.RightToLeft)
UA_check = QCheckBox("إلغاء تجميد النص العربي", OptionsWindow)
UA_check.setGeometry(QtCore.QRect(0, pos_y(4), checkbox_size[0], checkbox_size[1]))
UA_check.setLayoutDirection(QtCore.Qt.RightToLeft)
C_check = QCheckBox("تحويل النص", OptionsWindow)
C_check.setGeometry(QtCore.QRect(0, pos_y(5), checkbox_size[0], checkbox_size[1]))
C_check.setLayoutDirection(QtCore.Qt.RightToLeft)
UC_check = QCheckBox("إلغاء تحويل النص", OptionsWindow)
UC_check.setGeometry(QtCore.QRect(0, pos_y(6), checkbox_size[0], checkbox_size[1]))
UC_check.setLayoutDirection(QtCore.Qt.RightToLeft)
RT_check = QCheckBox("عكس النص", OptionsWindow)
RT_check.setGeometry(QtCore.QRect(0, pos_y(7), checkbox_size[0], checkbox_size[1]))
RT_check.setLayoutDirection(QtCore.Qt.RightToLeft)
RAO_check = QCheckBox("عكس العربية في النص (تجريبي)", OptionsWindow)
RAO_check.setGeometry(QtCore.QRect(0, pos_y(8), checkbox_size[0], checkbox_size[1]))
RAO_check.setLayoutDirection(QtCore.Qt.RightToLeft)
Ext_check = QCheckBox("استخرج من النص", OptionsWindow)
Ext_check.setGeometry(QtCore.QRect(0, pos_y(9), checkbox_size[0], checkbox_size[1]))
Ext_check.setLayoutDirection(QtCore.Qt.RightToLeft)
CB_check = QCheckBox("تحويل البايتات", OptionsWindow)
CB_check.setGeometry(QtCore.QRect(0, pos_y(10), checkbox_size[0], checkbox_size[1]))
CB_check.setLayoutDirection(QtCore.Qt.RightToLeft)
DH_check = QCheckBox("حذف الحركات المتتالية", OptionsWindow)
DH_check.setGeometry(QtCore.QRect(0, pos_y(11), checkbox_size[0], checkbox_size[1]))
DH_check.setLayoutDirection(QtCore.Qt.RightToLeft)

UC_database_button = QPushButton(OptionsWindow)
UC_database_button.setGeometry(QtCore.QRect(20, 310, 93, 56))
UC_database_button.setText("قاعدة بيانات\nالتحويل")

start_command = QTextEdit(OptionsWindow)
start_command.setGeometry(QtCore.QRect(10, 10, 70, 26))
start_com_label = QLabel(OptionsWindow)
start_com_label.setGeometry(QtCore.QRect(120, 10, 60, 26))
start_com_label.setText("قبل الأوامر:")
end_command = QTextEdit(OptionsWindow)
end_command.setGeometry(QtCore.QRect(10, 43, 70, 26))
end_com_end_label = QLabel(OptionsWindow)
end_com_end_label.setGeometry(QtCore.QRect(145, 43, 35, 26))
end_com_end_label.setText("بعدها:")
before_text_convert = QTextEdit(OptionsWindow)
before_text_convert.setGeometry(QtCore.QRect(10, 76, 70, 26))
before_text_convert_label = QLabel(OptionsWindow)
before_text_convert_label.setGeometry(QtCore.QRect(85, 76, 95, 26))
before_text_convert_label.setText("ما قبل النصوص:")
after_text_convert = QTextEdit(OptionsWindow)
after_text_convert.setGeometry(QtCore.QRect(10, 109, 70, 26))
after_text_convert_label = QLabel(OptionsWindow)
after_text_convert_label.setGeometry(QtCore.QRect(85, 109, 95, 26))
after_text_convert_label.setText("ما بعدها:")
min_text_convert = QTextEdit(OptionsWindow)
min_text_convert.setGeometry(QtCore.QRect(10, 142, 70, 26))
min_text_convert_label = QLabel(OptionsWindow)
min_text_convert_label.setGeometry(QtCore.QRect(85, 142, 95, 26))
min_text_convert_label.setText("أقصى حد لقصرها:")
max_text_convert = QTextEdit(OptionsWindow)
max_text_convert.setGeometry(QtCore.QRect(10, 175, 70, 26))
max_text_convert_label = QLabel(OptionsWindow)
max_text_convert_label.setGeometry(QtCore.QRect(85, 175, 95, 26))
max_text_convert_label.setText("أقصى حد لطولها:")
converted_byte = QTextEdit(OptionsWindow)
converted_byte.setGeometry(QtCore.QRect(10, 208, 70, 26))
converted_byte.setText(r'\xXY')
converted_byte_label = QLabel(OptionsWindow)
converted_byte_label.setGeometry(QtCore.QRect(85, 208, 95, 26))
converted_byte_label.setText("صيغة البايت المحول:")

Slash_check = QCheckBox(r"مراعاة: n, \t, \r, \a\ ", OptionsWindow)
Slash_check.setGeometry(QtCore.QRect(35, 235, 140, 26))
Slash_check.setLayoutDirection(QtCore.Qt.RightToLeft)
##


#النافذة الرئيسية
CMainWindow = QMainWindow()
CMainWindow.setFixedSize(685, 290)

label = QLabel(CMainWindow)
label.setGeometry(QtCore.QRect(580, 5, 81, 20))
label.setFont(labels_font)
label.setText("النص الداخل:")
label_2 = QLabel(CMainWindow)
label_2.setGeometry(QtCore.QRect(185, 5, 81, 20))
label_2.setFont(labels_font)
label_2.setText("النص الناتج:")

result_text = QTextEdit(CMainWindow)
result_text.setGeometry(QtCore.QRect(10, 30, 270, 250))
result_text.setFont(textbox_font)
entered_text = QTextEdit(CMainWindow)
entered_text.setGeometry(QtCore.QRect(405, 30, 270, 250))
entered_text.setFont(textbox_font)

convert_button = QPushButton(CMainWindow)
convert_button.setGeometry(QtCore.QRect(295, 95, 93, 28))
convert_button.setText("تحويل")
openfile_button = QPushButton(CMainWindow)
openfile_button.setGeometry(QtCore.QRect(295, 130, 93, 41))
openfile_button.setText("فتح ملف\nنص")
convert_files_button = QPushButton(CMainWindow)
convert_files_button.setGeometry(QtCore.QRect(295, 177, 93, 41))
convert_files_button.setText("فتح مجلد\nوتحويل ملفاته")


#نافذة الإدخال
EnteringWindow = QMainWindow()
EnteringWindow.setFixedSize(756, 330)

translate_text = QTextEdit(EnteringWindow)
translate_text.setGeometry(QtCore.QRect(13, 35, 301, 140))
translate_text.setFont(textbox_font)
original_text = QTextEdit(EnteringWindow)
original_text.setGeometry(QtCore.QRect(440, 35, 301, 140))
original_text.setFont(textbox_font)

label = QLabel(EnteringWindow)
label.setGeometry(QtCore.QRect(654, 10, 81, 20))
label.setFont(labels_font)
label.setText("النص الأصلي:")
label_2 = QLabel(EnteringWindow)
label_2.setGeometry(QtCore.QRect(220, 10, 81, 20))
label_2.setFont(labels_font)
label_2.setText("الترجمة:")

enter_button = QPushButton(EnteringWindow)
enter_button.setGeometry(QtCore.QRect(330, 45, 93, 28))
enter_button.setText("إدخال")
convert_enter_button = QPushButton(EnteringWindow)
convert_enter_button.setGeometry(QtCore.QRect(330, 80, 93, 41))
convert_enter_button.setText("تحويل\nوإدخال")
extract_button = QPushButton(EnteringWindow)
extract_button.setGeometry(QtCore.QRect(330, 130, 93, 28))
extract_button.setText("استخراج")

before_text = QTextEdit(EnteringWindow)
before_text.setGeometry(QtCore.QRect(25, 190, 50, 26))
before_text_label = QLabel(EnteringWindow)
before_text_label.setGeometry(QtCore.QRect(70, 190, 60, 26))
before_text_label.setText("ما يسبقه:")
after_text = QTextEdit(EnteringWindow)
after_text.setGeometry(QtCore.QRect(145, 190, 50, 26))
after_text_label = QLabel(EnteringWindow)
after_text_label.setGeometry(QtCore.QRect(215, 190, 160, 26))
after_text_label.setText("ما يلحق النص في الملفات:")

min_text = QTextEdit(EnteringWindow)
min_text.setGeometry(QtCore.QRect(145, 223, 50, 26))
min_text_label = QLabel(EnteringWindow)
min_text_label.setGeometry(QtCore.QRect(175, 223, 200, 26))
min_text_label.setText("أقصى حد لقصر النصوص المستخرجة:")
max_text = QTextEdit(EnteringWindow)
max_text.setGeometry(QtCore.QRect(25, 223, 50, 26))
max_text_label = QLabel(EnteringWindow)
max_text_label.setGeometry(QtCore.QRect(70, 223, 60, 26))
max_text_label.setText("وطولها:")

database_check = QCheckBox("استخدام قاعدة بيانات النصوص للإدخال", EnteringWindow)
database_check.setGeometry(QtCore.QRect(520, 193, 200, 16))
database_check.setLayoutDirection(QtCore.Qt.RightToLeft)
too_long_check = QCheckBox("عدم إدخال ترجمات أطول من النص الأصلي (بقيم الهيكس)", EnteringWindow)
too_long_check.setGeometry(QtCore.QRect(440, 218,280, 16))
too_long_check.setLayoutDirection(QtCore.Qt.RightToLeft)
translation_place_check = QCheckBox(":مكان الترجمة في حال كانت أقصر", EnteringWindow)
translation_place_check.setGeometry(QtCore.QRect(470, 243,250, 16))
translation_place_check.setLayoutDirection(QtCore.Qt.RightToLeft)

first_radio = QRadioButton(EnteringWindow)
first_radio.setGeometry(QtCore.QRect(580, 268,100, 16))
first_radio.setText("أول")
first_radio.setLayoutDirection(QtCore.Qt.RightToLeft)
middle_radio = QRadioButton(EnteringWindow)
middle_radio.setGeometry(QtCore.QRect(510, 268,100, 16))
middle_radio.setText("وسط")
middle_radio.setLayoutDirection(QtCore.Qt.RightToLeft)
last_radio = QRadioButton(EnteringWindow)
last_radio.setGeometry(QtCore.QRect(440, 268,100, 16))
last_radio.setText("آخر")
last_radio.setLayoutDirection(QtCore.Qt.RightToLeft)

input_from_folder = QPushButton(EnteringWindow)
input_from_folder.setGeometry(QtCore.QRect(335, 270, 93, 41))
input_from_folder.setText("المجلد الحاوي\nللملفات")
output_from_folder = QPushButton(EnteringWindow)
output_from_folder.setGeometry(QtCore.QRect(230, 270, 93, 41))
output_from_folder.setText("مجلد الملفات\nبعد الإدخال")
text_database_button = QPushButton(EnteringWindow)
text_database_button.setGeometry(QtCore.QRect(125, 270, 93, 41))
text_database_button.setText("فتح قاعدة\nبيانات النصوص")
extract_database_button = QPushButton(EnteringWindow)
extract_database_button.setGeometry(QtCore.QRect(20, 270, 93, 41))
extract_database_button.setText("فتح قاعدة\nبيانات الاستخراج")

#توصيل الإشارات
convert_button.clicked.connect(lambda: result_text.setPlainText(convert(entered_text.toPlainText())))
openfile_button.clicked.connect(lambda: open_def(4))
convert_files_button.clicked.connect(lambda:convertFiles())

text_database_button.clicked.connect(lambda: open_def(0))
UC_database_button.clicked.connect(lambda: open_def(1))
extract_database_button.clicked.connect(lambda: open_def(3))

enter_button.clicked.connect(lambda: enter(False))
extract_button.clicked.connect(lambda: extract())
convert_enter_button.clicked.connect(lambda: enter())
input_from_folder.clicked.connect(lambda: open_def(5))
output_from_folder.clicked.connect(lambda: open_def(6))


#المتغيرات
converting_database_directory = r'Parts/Scripts/Un-Converting_Database.ate'
chars_width_database_directory = r'Parts/Scripts/Chars_Width_Database.ate'
text_database_directory = r'OtherFiles/TextTable.xlsx'
extracted_text_database_directory = r'OtherFiles/ExtractedTextTable.xlsx'
input_folder, output_folder = r'OtherFiles/_FilesFolder/', r'OtherFiles/_AfterEnteringFolder/'
cell_byte = True

if path.exists(converting_database_directory): convert_database = Take_From_Table(converting_database_directory)
else: convert_database = {}

#الدوال
def dir_list(path): return [root+'/'+'{}{}'.format('', f) for root, dirs, files in walk(path) for f in files]

def cell_bytes():
    if '[b]' in start_command.toPlainText(): cell_bytes._start_command = bytearray.fromhex(start_command.toPlainText().replace('[b]', '')).decode()
    else: cell_bytes._start_command = start_command.toPlainText()
    if '[b]' in end_command.toPlainText(): cell_bytes._end_command = bytearray.fromhex(end_command.toPlainText().replace('[b]', '')).decode()
    else: cell_bytes._end_command = end_command.toPlainText()
    if '[b]' in before_text_convert.toPlainText(): cell_bytes._before_text_convert = bytearray.fromhex(before_text_convert.toPlainText().replace('[b]', '')).decode()
    else: cell_bytes._before_text_convert = before_text_convert.toPlainText()
    if '[b]' in after_text_convert.toPlainText(): cell_bytes._after_text_convert = bytearray.fromhex(after_text_convert.toPlainText().replace('[b]', '')).decode()
    else: cell_bytes._after_text_convert = after_text_convert.toPlainText()
    if '[b]' in converted_byte.toPlainText(): cell_bytes._converted_byte = bytearray.fromhex(converted_byte.toPlainText().replace('[b]', '')).decode()
    else: cell_bytes._converted_byte = converted_byte.toPlainText()

    if Slash_check.isChecked():
        cell_bytes._start_command = cell_bytes._start_command.replace(r'\n', '\n').replace(r'\t', '\t').replace(r'\r', '\r').replace(r'\a', '\a')
        cell_bytes._end_command = cell_bytes._end_command.replace(r'\n', '\n').replace(r'\t', '\t').replace(r'\r', '\r').replace(r'\a', '\a')
        cell_bytes._before_text_convert = cell_bytes._before_text_convert.replace(r'\n', '\n').replace(r'\t', '\t').replace(r'\r', '\r').replace(r'\a', '\a')
        cell_bytes._after_text_convert = cell_bytes._after_text_convert.replace(r'\n', '\n').replace(r'\t', '\t').replace(r'\r', '\r').replace(r'\a', '\a')
        cell_bytes._converted_byte = cell_bytes._converted_byte.replace(r'\n', '\n').replace(r'\t', '\t').replace(r'\r', '\r').replace(r'\a', '\a')

def open_def(num):
    if num == 0:
        fileName, _ = QFileDialog.getOpenFileName(EnteringWindow, 'قاعدة بيانات النص', '' , '*.xlsx')
        if path.exists(fileName) and fileName != '/':
            global text_database_directory
            text_database_directory = fileName
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 1:
        fileName, _ = QFileDialog.getOpenFileName(OptionsWindow, 'قاعدة بيانات التحويل', '' , '*.ate')
        if path.exists(fileName) and fileName != '/':
            global converting_database_directory, convert_database
            converting_database_directory = fileName
            convert_database = Take_From_Table(fileName)
            QMessageBox.about(OptionsWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 3:
        fileName, _ = QFileDialog.getOpenFileName(EnteringWindow, 'قاعدة بيانات الاستخراج', '' , '*.ate')
        if path.exists(fileName) and fileName != '/':
            global extracted_text_database_directory
            extracted_text_database_directory = fileName
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 4:
        fileName, _ = QFileDialog.getOpenFileName(CMainWindow, 'ملف نص', '' , '*')
        if path.exists(fileName) and fileName != '/':
            entered_text.setPlainText(open(fileName, 'r', encoding='utf-8').read())
            QMessageBox.about(CMainWindow, "!!تهانيّ", "تم اختيار ملف النص.")
    elif num == 5:
        folder = str(QFileDialog.getExistingDirectory(EnteringWindow, "Select Directory"))+'/'
        if path.exists(folder) and folder != '/':
            global input_folder
            input_folder = folder
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")
    elif num == 6:
        folder = str(QFileDialog.getExistingDirectory(EnteringWindow, "Select Directory"))+'/'
        if path.exists(folder) and folder != '/':
            global output_folder
            output_folder = folder
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")

def convert(text):
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if text == '': return
    if (C_check.isChecked() or UC_check.isChecked()) and not path.exists(converting_database_directory):
        QMessageBox.about(CMainWindow, "!!خطأ", "قاعدة بيانات التحويل غير موجودة")
    
    if cell_byte: cell_bytes()
    
    if Ext_check.isChecked():#Extract from text
        if cell_bytes._before_text_convert == '' or cell_bytes._after_text_convert == '':
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما قبل النصوص، ما بعدها.\nعلى الأقل للاستخراج.")
            return
        
        mini = min_text_convert.toPlainText()
        maxi = max_text_convert.toPlainText()
        if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode()
        if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode()
        if mini == '': mini = 0
        else: mini = int(mini)
        if maxi == '': maxi = 0
        else: maxi = int(maxi)
        
        if mini > maxi:
            QMessageBox.about(EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
            return
        
        text = Extract(text, cell_bytes._before_text_convert, cell_bytes._after_text_convert, mini, maxi)
        text = '\n'.join(text)
    
    if DH_check.isChecked(): text = handle_harakat(text)#Delete Harakat
    if DDL_check.isChecked(): text = DDL(text)#Delete Duplicated lines
    if SSL_check.isChecked(): text = Sort(text)#Sort short to long
    if SLS_check.isChecked(): text = Sort(text, False)#Sort long to short
    if RA_check.isChecked() or C_check.isChecked(): text = Un_Freeze(text)#Freeze Arabic
    if C_check.isChecked(): text = Convert(text, convert_database, True, cell_bytes._start_command, cell_bytes._end_command)#Convert
    if UC_check.isChecked(): text = Convert(text, convert_database, False, cell_bytes._start_command, cell_bytes._end_command)#Unconvert
    if UA_check.isChecked() or UC_check.isChecked(): text = Un_Freeze(text, False)#UnFreeze Arabic
    if CB_check.isChecked(): text = convert_bytes(text, cell_bytes._converted_byte)#Convert bytes
    if RT_check.isChecked(): text = Reverse(text, cell_bytes._start_command, cell_bytes._end_command, '\n\n', '\n')#Reverse whole text
    if RAO_check.isChecked(): text = Reverse(text, cell_bytes._start_command, cell_bytes._end_command, '\n\n', '\n', False)#‫Reverse Arabic only
    return text

def convertFiles():
    folder = str(QFileDialog.getExistingDirectory(CMainWindow, "اختر مجلداً"))+'/'
    files = dir_list(folder)
    
    for file in files:
        content = open(file, 'r', encoding='utf-8').read()
        if not content: continue
        open(file, 'w', encoding='utf-8').write(str(convert(content)))
    
    QMessageBox.about(EnteringWindow, "!!تهانينا", "انتهى تحويل الملفات.")

def enter(convert_bool = True):
    ##المتغيرات
    textList = []
    before = before_text.toPlainText()
    after = after_text.toPlainText()
    if '[b]' in before: before = bytearray.fromhex(before.replace('[b]', '')).decode() 
    if '[b]' in after: after = bytearray.fromhex(after.replace('[b]', '')).decode() 
    
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if C_check.isChecked() or UC_check.isChecked():
        if not path.exists(converting_database_directory):
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات التحويل غير موجودة.")
            return
    if not path.exists(input_folder):
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nالمجلد الحاوي للملفات غير موجود.")
        return
    if not path.exists(output_folder):
        mkdir(output_folder)
    files_list = dir_list('./'+input_folder)
    if len(files_list) == 0:
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nلا توجد أي ملفات للإدخال إليها.")
        return
    
    
    if database_check.isChecked():
        if not path.exists(text_database_directory):
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات النصوص غير موجودة.")
            return
        text_xlsx = openpyxl.load_workbook(text_database_directory)
        text_table = text_xlsx.get_sheet_by_name("Main")
        for cell in range(2, len(text_table['A'])+1):
            original_cell_value = text_table['A'+str(cell)].value
            translate_cell_value = text_table['B'+str(cell)].value
            
            if original_cell_value and translate_cell_value:
                textList.append([original_cell_value, translate_cell_value])
        
        sorted(textList, key=lambda x: len(str(x[0])), reverse=True)
    else:
        if original_text.toPlainText(): return
        textList.append([original_text.toPlainText(), translate_text.toPlainText()])
    
    for filename in files_list:
        with open(filename, 'rb') as f:
            file_content = f.read()
        
        for i in range(len(textList)):
            text, translation = textList[i][0], textList[i][1]
            text = before + text + after
            translation = before + translation + after
            
            if translation_place_check.isChecked() and len(translation.encode('utf-8').hex()) < len(text.encode('utf-8').hex()):
                spaces_count = (len(text.encode('utf-8').hex()) // 2) - (len(translation.encode('utf-8').hex()) // 2)
                if first_radio.isChecked():#first
                    for i in range(spaces_count):
                        translation += ' '
                elif middle_radio.isChecked():#middle
                    for i in range(spaces_count):
                        if i % 2 == 0:
                            translation += ' '
                        else:
                            translation = ' ' + translation
                elif last_radio.isChecked():#last
                    for i in range(spaces_count):
                        translation = ' ' + translation
            
            if convert_bool: translation = convert(translation)
            if too_long_check.isChecked() and len(translation.encode('utf-8').hex()) > len(text.encode('utf-8').hex()): continue
            file_content = file_content.replace(bytes(text, 'utf-8'), bytes(translation, 'utf-8'), 1)
        
        directory = filename.replace('./'+input_folder, output_folder)
        makedirs(path.dirname(directory), exist_ok=True)
        open(directory, 'wb').write(file_content)
    
    QMessageBox.about(EnteringWindow, "!!تهانيّ", "انتهى الإدخال.")

def extract():
    before = before_text.toPlainText()
    after = after_text.toPlainText()
    if '[b]' in before: before = bytearray.fromhex(before.replace('[b]', '')).decode() 
    if '[b]' in after: after = bytearray.fromhex(after.replace('[b]', '')).decode() 
    if before == '' or after == '':
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما يسبق النصوص، ما يلحقها.\nعلى الأقل.")
        return
    files_list = dir_list(input_folder)
    if len(files_list) == 0:
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nلا توجد أي ملفات للاستخراج منها.")
        return
    
    mini = min_text.toPlainText()
    maxi = max_text.toPlainText()
    if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode() 
    if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode() 
    
    if mini: mini = int(mini)
    else: mini = 0
    if maxi: maxi = int(maxi)
    else: maxi = 0
    
    if mini > maxi:
        QMessageBox.about(EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
        return
    
    extracted_xlsx = openpyxl.load_workbook(extracted_text_database_directory)
    sheet = extracted_xlsx.get_sheet_by_name("Main")
    row = 2
    '''
    def put_in_sheet(text):
        print(row)
        print(text)
        row += 1
        print(row)
    '''
    sheet.delete_cols(1, 2)
    sheet['A1'].value = "النص الأصلي"
    sheet['A1'].font = Font(bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet['A1'].fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
    
    for filename in files_list:
        with open(filename, 'r', encoding="cp437") as f:
            file_content = f.read()
        
        extracted = Extract(file_content, before, after, mini, maxi)
        
        if len(extracted):
            sheet['A'+str(row)].value = filename
            sheet['A'+str(row)].font = Font(bold=True)
            sheet['A'+str(row)].alignment = Alignment(vertical='center', wrap_text=True)
            sheet['A'+str(row)].fill = PatternFill(fill_type='solid', start_color='D112D1', end_color='D112D1')
            row += 1
            
            '''map(put_in_sheet, extracted)'''
            for item in extracted:
                sheet['A'+str(row)].font = Font(bold=False)
                sheet['A'+str(row)].value = item
                row += 1
    
    extracted_xlsx.save(extracted_text_database_directory)
    QMessageBox.about(EnteringWindow, "!!تهانيّ", "انتهى الاستخراج.")